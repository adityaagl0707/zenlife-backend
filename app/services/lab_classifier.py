"""Lab result parser and severity classifier for ZenLife."""
import re
from typing import Optional


# Master marker database — all 121 markers from the ZenLife panel
MARKERS = [
    {"name": "LDL Cholesterol", "description": "Bad Cholesterol", "normal_range": "< 100", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Fasting Blood Glucose", "description": "Fasting Blood Sugar", "normal_range": "70 - 100", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Homocysteine", "description": "Heart & Brain Health Indicator", "normal_range": "< 15", "unit": "µmol/L", "test_type": "blood_urine", "organs": ["Heart", "Brain & Cognitive Health"]},
    {"name": "Total Cholesterol", "description": "Cholesterol Sum", "normal_range": "< 200", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "HDL Cholesterol", "description": "Good Cholesterol", "normal_range": "> 50", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Triglycerides", "description": "Circulating Blood Fat", "normal_range": "< 150", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart", "Endocrine & Metabolic Health"]},
    {"name": "Non-HDL Cholesterol", "description": "All Bad Cholesterol", "normal_range": "< 130", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "AST", "description": "Liver & Heart Enzyme", "normal_range": "< 31", "unit": "U/L", "test_type": "blood_urine", "organs": ["Heart", "Liver & Digestive Health"]},
    {"name": "Total Protein", "description": "Nutritional & Liver Health Marker", "normal_range": "5.7 - 8.2", "unit": "g/dL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "General Health, Blood and Nutrients"]},
    {"name": "Globulin", "description": "Immune & Protein Balance", "normal_range": "2.5 - 3.4", "unit": "gm/dL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "ALT", "description": "Liver Health Enzyme", "normal_range": "< 34", "unit": "U/L", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "Creatinine", "description": "Kidney Filtration Marker", "normal_range": "0.55 - 1.02", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Uric Acid", "description": "Kidney Stone Risk", "normal_range": "3.2 - 6.1", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health", "Bone, Muscle & Joint Health"]},
    {"name": "Basophils", "description": "Allergy Response Cells", "normal_range": "0.02 - 0.10", "unit": "%", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Monocytes", "description": "Inflammation & Repair", "normal_range": "0.20 - 1.00", "unit": "%", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Lymphocytes", "description": "Viral Infection Defense", "normal_range": "20 - 40", "unit": "%", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Eosinophils", "description": "Allergy & Parasite Defense", "normal_range": "0.02 - 0.50", "unit": "%", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Vitamin D", "description": "Bone & Immune Health", "normal_range": "30 - 100", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health", "Bone, Muscle & Joint Health"]},
    {"name": "Copper", "description": "Red Blood Cell Support", "normal_range": "80 - 155", "unit": "µg/L", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Cortisol", "description": "Stress Hormone", "normal_range": "5 - 23", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health", "Brain & Cognitive Health"]},
    {"name": "Folate (B9)", "description": "Cell Growth & Repair", "normal_range": "> 5.38", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Brain & Cognitive Health", "General Health, Blood and Nutrients"]},
    {"name": "DHEA", "description": "Neurosteroid & Cognitive Health", "normal_range": "106 - 464", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health", "Reproductive Health"]},
    {"name": "IgE", "description": "Immune Response & Allergies", "normal_range": "< 158", "unit": "IU/mL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health", "Lung & Respiratory Health"]},
    {"name": "VLDL Cholesterol", "description": "Triglyceride Transporter", "normal_range": "5 - 40", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Triglycerides / HDL", "description": "Metabolic Health Marker", "normal_range": "< 3.12", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Heart", "Endocrine & Metabolic Health"]},
    {"name": "Total Cholesterol / HDL", "description": "Heart Risk Indicator", "normal_range": "3 - 5", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "HDL/LDL Ratio", "description": "Cholesterol Risk Indicator", "normal_range": "> 0.40", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Urine Albumin/Creatinine", "description": "Kidney Function Marker", "normal_range": "< 30", "unit": "µg/mg", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Iron % Saturation", "description": "Iron Utilization Rate", "normal_range": "13 - 45", "unit": "%", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "RBC", "description": "Red Blood Cell Count", "normal_range": "3.80 - 4.80", "unit": "x10⁶/µL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "TPO", "description": "Autoimmune Thyroid Marker", "normal_range": "< 9", "unit": "IU/mL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Calcium", "description": "Mineral Balance Marker", "normal_range": "8.8 - 10.6", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health", "Bone, Muscle & Joint Health"]},
    {"name": "GGT", "description": "Liver & Bile Duct Health", "normal_range": "< 38", "unit": "U/L", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "BUN", "description": "Kidney Waste Filtration", "normal_range": "7.94 - 20.07", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Sodium", "description": "Fluid Balance Marker", "normal_range": "136 - 145", "unit": "mmol/L", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Potassium", "description": "Electrolyte Balance", "normal_range": "3.5 - 5.1", "unit": "mmol/L", "test_type": "blood_urine", "organs": ["Heart", "Kidney & Urinary Health"]},
    {"name": "BUN / Creatinine Ratio", "description": "Kidney Function & Hydration", "normal_range": "9 - 23", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Chloride", "description": "Fluid Balance Marker", "normal_range": "98 - 107", "unit": "mmol/L", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "eGFR", "description": "Kidney Filtration Score", "normal_range": "> 90", "unit": "mL/min/1.73m²", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Leukocytes", "description": "Infection Marker", "normal_range": "4 - 10", "unit": "x10³/µL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Iron", "description": "Oxygen Transport", "normal_range": "50 - 170", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "TIBC", "description": "Iron Binding Capacity", "normal_range": "215 - 535", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Vitamin A", "description": "Vision & Immunity", "normal_range": "300 - 800", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health", "General Health, Blood and Nutrients"]},
    {"name": "Zinc", "description": "Immunity & Healing", "normal_range": "4000 - 9000", "unit": "µg/L", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health", "General Health, Blood and Nutrients"]},
    {"name": "Selenium", "description": "Antioxidant Defense", "normal_range": "60 - 340", "unit": "µg/L", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health", "General Health, Blood and Nutrients"]},
    {"name": "WBC", "description": "Immune Cell Count", "normal_range": "4 - 10", "unit": "x10³/µL", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "MPV", "description": "Platelet Size Indicator", "normal_range": "6.5 - 12", "unit": "fL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Hemoglobin", "description": "Oxygen Transport", "normal_range": "12 - 15", "unit": "g/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Platelet Count", "description": "Clotting Ability", "normal_range": "150 - 410", "unit": "x10³/µL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Hematocrit", "description": "Blood Cell Volume", "normal_range": "40 - 50", "unit": "%", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "MCV", "description": "Red Blood Cell Size", "normal_range": "83 - 101", "unit": "fL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "RDW", "description": "Red Cell Variation", "normal_range": "11.6 - 14", "unit": "%", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "MCH", "description": "Hemoglobin Per Cell", "normal_range": "27 - 32", "unit": "pg", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "MCHC", "description": "Hemoglobin Concentration", "normal_range": "31.5 - 34.5", "unit": "g/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "PSA", "description": "Prostate Health Marker", "normal_range": "< 4", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Reproductive Health"]},
    {"name": "UIBC", "description": "Unsaturated Iron-Binding Capacity", "normal_range": "162 - 368", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Urea/Creatinine Ratio", "description": "Kidney Function & Hydration", "normal_range": "< 52", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Urea (Calculated)", "description": "Kidney Function & Metabolism", "normal_range": "17 - 43", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "Bilirubin (Indirect)", "description": "Liver & RBC Health", "normal_range": "0 - 0.9", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "General Health, Blood and Nutrients"]},
    {"name": "Plateletcrit (PCT)", "description": "Platelet Volume", "normal_range": "0.19 - 0.39", "unit": "%", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Vitamin B12", "description": "Nerve & Blood Health", "normal_range": "197 - 771", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Brain & Cognitive Health", "General Health, Blood and Nutrients"]},
    {"name": "Albumin/Globulin Ratio", "description": "Liver & Immune Balance", "normal_range": "0.9 - 2.0", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "Inflammation & Immune Health"]},
    {"name": "Albumin", "description": "Blood Protein & Fluid Balance", "normal_range": "3.2 - 4.8", "unit": "g/dL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "General Health, Blood and Nutrients"]},
    {"name": "Total Bilirubin", "description": "Liver & Blood Health", "normal_range": "0.30 - 1.29", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "Bilirubin Direct", "description": "Liver Function & Bile Flow", "normal_range": "< 0.30", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "ALP", "description": "Liver & Bone Health Enzyme", "normal_range": "45 - 129", "unit": "U/L", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "Bone, Muscle & Joint Health"]},
    {"name": "Lipase", "description": "Fat Digestion Enzyme", "normal_range": "5.6 - 51.3", "unit": "U/L", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "Amylase", "description": "Carb Digestion Enzyme", "normal_range": "28 - 100", "unit": "U/L", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "AST/ALT Ratio", "description": "Liver Function Indicator", "normal_range": "< 2", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "Insulin", "description": "Insulin Sensitivity", "normal_range": "1.9 - 23", "unit": "µIU/mL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "HbA1c", "description": "Long-Term Blood Sugar", "normal_range": "< 5.7", "unit": "%", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "TSH", "description": "Thyroid Function Regulator", "normal_range": "0.30 - 5.50", "unit": "µIU/mL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "T4 Free", "description": "Thyroid Hormone Reserve", "normal_range": "0.93 - 1.70", "unit": "ng/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "T3 Free", "description": "Active Thyroid Hormone", "normal_range": "2 - 4.4", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Average Blood Glucose", "description": "Long-Term Sugar Control", "normal_range": "< 120", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Testosterone", "description": "Hormone for Vitality", "normal_range": "280 - 800", "unit": "ng/dL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health", "Reproductive Health"]},
    {"name": "hs-CRP", "description": "Inflammation Marker", "normal_range": "< 3", "unit": "mg/L", "test_type": "blood_urine", "organs": ["Heart", "Inflammation & Immune Health"]},
    {"name": "Neutrophils", "description": "Bacterial Infection Defense", "normal_range": "40 - 80", "unit": "%", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "Lipoprotein (a)", "description": "Genetic Heart Risk", "normal_range": "< 30", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Apolipoprotein B", "description": "Bad Cholesterol Carrier", "normal_range": "53 - 138", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "LDL/HDL Ratio", "description": "Cholesterol Balance Ratio", "normal_range": "1.5 - 3.5", "unit": "Ratio", "test_type": "blood_urine", "organs": ["Heart"]},
    {"name": "Ferritin", "description": "Iron Storage", "normal_range": "4.63 - 204", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients"]},
    {"name": "Magnesium, RBC", "description": "Cellular Magnesium", "normal_range": "1.9 - 3.1", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["General Health, Blood and Nutrients", "Bone, Muscle & Joint Health"]},
    {"name": "Cortisol (Urine)", "description": "Adrenal Function", "normal_range": "4 - 176", "unit": "µg/24h", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Agatston Score", "description": "Coronary Calcium Score", "normal_range": "< 100", "unit": "", "test_type": "calcium_score", "organs": ["Heart"]},
    {"name": "Visceral Fat Area", "description": "Abdominal Fat Measurement", "normal_range": "< 100", "unit": "cm²", "test_type": "dexa", "organs": ["Endocrine & Metabolic Health", "Heart"]},
    {"name": "Gynoid Fat", "description": "Hip/Thigh Fat %", "normal_range": "< 30", "unit": "%", "test_type": "dexa", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Total Body Fat %", "description": "Overall Body Fat", "normal_range": "10 - 25", "unit": "%", "test_type": "dexa", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Bone Mineral Density (Spine)", "description": "Spinal Bone Strength", "normal_range": "> -1.0", "unit": "T-score", "test_type": "dexa", "organs": ["Bone, Muscle & Joint Health"]},
    {"name": "Bone Mineral Density (Hip)", "description": "Hip Bone Strength", "normal_range": "> -1.0", "unit": "T-score", "test_type": "dexa", "organs": ["Bone, Muscle & Joint Health"]},
    {"name": "Lean Body Mass", "description": "Muscle Mass", "normal_range": "> 60", "unit": "%", "test_type": "dexa", "organs": ["Bone, Muscle & Joint Health"]},
    {"name": "Heart Rate (ECG)", "description": "Resting Heart Rate", "normal_range": "60 - 100", "unit": "bpm", "test_type": "ecg", "organs": ["Heart"]},
    {"name": "PR Interval", "description": "Heart Conduction", "normal_range": "120 - 200", "unit": "ms", "test_type": "ecg", "organs": ["Heart"]},
    {"name": "QRS Duration", "description": "Ventricular Conduction", "normal_range": "< 120", "unit": "ms", "test_type": "ecg", "organs": ["Heart"]},
    {"name": "QTc Interval", "description": "Heart Rhythm Safety", "normal_range": "< 450", "unit": "ms", "test_type": "ecg", "organs": ["Heart"]},
    # Advanced blood markers
    {"name": "ApoA1", "description": "Protective Lipoprotein Carrier", "normal_range": "120 - 160", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Heart", "Vascular Health"]},
    {"name": "C-Peptide", "description": "Insulin Reserve Marker", "normal_range": "0.8 - 3.5", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health"]},
    {"name": "Fibrinogen", "description": "Clotting & Inflammation Marker", "normal_range": "200 - 400", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Vascular Health", "Inflammation & Immune Health"]},
    {"name": "ESR", "description": "Inflammation Screen", "normal_range": "0 - 20", "unit": "mm/hr", "test_type": "blood_urine", "organs": ["Inflammation & Immune Health"]},
    {"name": "NT-proBNP", "description": "Heart Stress Marker", "normal_range": "< 125", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Heart", "Vascular Health"]},
    {"name": "Free Testosterone", "description": "Bioavailable Testosterone", "normal_range": "9 - 30", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health", "Reproductive Health"]},
    {"name": "SHBG", "description": "Sex Hormone Binding Globulin", "normal_range": "10 - 57", "unit": "nmol/L", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health"]},
    {"name": "Estradiol (E2)", "description": "Sex Hormone for Bone & Heart", "normal_range": "20 - 150", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health", "Reproductive Health"]},
    {"name": "IGF-1", "description": "Growth Hormone Axis Marker", "normal_range": "100 - 250", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health", "Endocrine & Metabolic Health"]},
    {"name": "Cystatin C", "description": "Precise Kidney Filtration Marker", "normal_range": "0.5 - 1.0", "unit": "mg/L", "test_type": "blood_urine", "organs": ["Kidney & Urinary Health"]},
    {"name": "PTH", "description": "Parathyroid Hormone", "normal_range": "15 - 65", "unit": "pg/mL", "test_type": "blood_urine", "organs": ["Bone, Muscle & Joint Health", "Kidney & Urinary Health"]},
    {"name": "Phosphorus", "description": "Bone & Energy Mineral", "normal_range": "2.5 - 4.5", "unit": "mg/dL", "test_type": "blood_urine", "organs": ["Bone, Muscle & Joint Health", "General Health, Blood & Nutrients"]},
    {"name": "Vitamin E", "description": "Antioxidant Defense", "normal_range": "5.5 - 17", "unit": "mg/L", "test_type": "blood_urine", "organs": ["General Health, Blood & Nutrients", "Inflammation & Immune Health"]},
    {"name": "Reverse T3", "description": "Inactive Thyroid Hormone", "normal_range": "10 - 24", "unit": "ng/dL", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health", "Endocrine & Metabolic Health"]},
    {"name": "LDH", "description": "Tissue Damage Marker", "normal_range": "140 - 280", "unit": "U/L", "test_type": "blood_urine", "organs": ["General Health, Blood & Nutrients"]},
    {"name": "D-Dimer", "description": "Clotting Activity Marker", "normal_range": "< 0.5", "unit": "µg/mL", "test_type": "blood_urine", "organs": ["Vascular Health", "Inflammation & Immune Health"]},
    {"name": "PT / INR", "description": "Blood Clotting Speed", "normal_range": "0.8 - 1.1", "unit": "INR", "test_type": "blood_urine", "organs": ["Vascular Health", "Liver & Digestive Health"]},
    {"name": "Anti-TG", "description": "Thyroid Autoimmune Marker", "normal_range": "< 115", "unit": "IU/mL", "test_type": "blood_urine", "organs": ["Hormonal & Vitality Health", "Inflammation & Immune Health"]},
    {"name": "AFP", "description": "Liver Cancer Screening Marker", "normal_range": "< 10", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health"]},
    {"name": "CEA", "description": "Colorectal Cancer Marker", "normal_range": "< 3", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Liver & Digestive Health", "General Health, Blood & Nutrients"]},
    {"name": "CA-125", "description": "Ovarian Cancer Marker", "normal_range": "< 35", "unit": "U/mL", "test_type": "blood_urine", "organs": ["Reproductive Health"]},
    {"name": "HOMA-IR", "description": "Insulin Resistance Score", "normal_range": "< 1.5", "unit": "index", "test_type": "blood_urine", "organs": ["Endocrine & Metabolic Health", "Vascular Health"]},
    # Imaging markers
    {"name": "Carotid CIMT", "description": "Artery Wall Thickness / Vascular Age", "normal_range": "< 0.9", "unit": "mm", "test_type": "usg", "organs": ["Vascular Health", "Heart"]},
    {"name": "Carotid Plaque Score", "description": "Arterial Plaque Burden", "normal_range": "0", "unit": "score", "test_type": "usg", "organs": ["Vascular Health", "Heart"]},
    {"name": "Liver Steatosis Grade", "description": "Fatty Liver Severity", "normal_range": "Grade 0", "unit": "", "test_type": "usg", "organs": ["Liver & Digestive Health", "Endocrine & Metabolic Health"]},
    {"name": "Renal Cortical Thickness", "description": "Kidney Reserve Marker", "normal_range": "> 10", "unit": "mm", "test_type": "usg", "organs": ["Kidney & Urinary Health"]},
    {"name": "Liver Fat %", "description": "MRI-measured Liver Fat (PDFF)", "normal_range": "< 5", "unit": "%", "test_type": "mri", "organs": ["Liver & Digestive Health", "Endocrine & Metabolic Health"]},
    {"name": "White Matter Hyperintensity", "description": "Brain Small Vessel Disease Burden", "normal_range": "None", "unit": "", "test_type": "mri", "organs": ["Mental & Stress Resilience", "Brain & Cognitive Health"]},
    {"name": "Prostate Volume", "description": "Prostate Size", "normal_range": "< 30", "unit": "cc", "test_type": "mri", "organs": ["Reproductive Health"]},
    {"name": "ASMI", "description": "Appendicular Skeletal Muscle Index", "normal_range": "> 7.0", "unit": "kg/m²", "test_type": "dexa", "organs": ["Bone, Muscle & Joint Health"]},
    {"name": "Fat Mass Index", "description": "Standardised Body Fat", "normal_range": "< 9", "unit": "kg/m²", "test_type": "dexa", "organs": ["Endocrine & Metabolic Health", "Vascular Health"]},
    {"name": "Trunk:Limb Fat Ratio", "description": "Central vs Peripheral Fat", "normal_range": "< 1.2", "unit": "ratio", "test_type": "dexa", "organs": ["Endocrine & Metabolic Health", "Vascular Health"]},
    {"name": "P-Wave Duration", "description": "Atrial Conduction Time / AF Risk", "normal_range": "< 120", "unit": "ms", "test_type": "ecg", "organs": ["Heart", "Vascular Health"]},
    # Female-specific markers
    {"name": "FSH", "description": "Follicle Stimulating Hormone", "normal_range": "3 - 10", "unit": "mIU/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "LH", "description": "Luteinizing Hormone", "normal_range": "2 - 15", "unit": "mIU/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "Progesterone", "description": "Female Reproductive Hormone", "normal_range": "0.2 - 1.5", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "Prolactin", "description": "Lactation & Reproductive Hormone", "normal_range": "2 - 29", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "AMH", "description": "Anti-Müllerian Hormone / Ovarian Reserve", "normal_range": "1.0 - 3.5", "unit": "ng/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "HE4", "description": "Human Epididymis Protein 4 / Ovarian Cancer Marker", "normal_range": "< 70", "unit": "pmol/L", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "CA 15-3", "description": "Breast Cancer Marker", "normal_range": "< 25", "unit": "U/mL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    {"name": "DHEA-S", "description": "Adrenal Androgen (Female)", "normal_range": "35 - 430", "unit": "µg/dL", "test_type": "blood_urine", "organs": ["Women's Health"]},
    # Female-specific imaging markers
    {"name": "Breast Ultrasound", "description": "Breast Tissue Assessment", "normal_range": "Normal", "unit": "", "test_type": "usg", "organs": ["Women's Health"]},
    {"name": "Transvaginal Ultrasound", "description": "Pelvic Organ Assessment", "normal_range": "Normal", "unit": "", "test_type": "usg", "organs": ["Women's Health"]},
    {"name": "Endometrial Thickness", "description": "Uterine Lining Thickness", "normal_range": "< 12", "unit": "mm", "test_type": "usg", "organs": ["Women's Health"]},
    {"name": "Mammography", "description": "Breast Cancer Screening", "normal_range": "BIRADS 1", "unit": "", "test_type": "usg", "organs": ["Women's Health"]},
]


def _extract_number(value_str: str) -> Optional[float]:
    """Extract the first numeric value from a string like '116 mg/dL' → 116.0"""
    if not value_str:
        return None
    match = re.search(r'[-+]?\d*\.?\d+', str(value_str).replace(',', ''))
    return float(match.group()) if match else None


def _parse_range(range_str: str):
    """
    Parse a normal range string into (lower, upper, is_lower_only, is_upper_only).
    Returns (lower, upper) where None means no bound.
    """
    s = str(range_str).strip()
    # "< X" or "<= X"
    m = re.match(r'^[<≤]=?\s*([\d.]+)', s)
    if m:
        return (None, float(m.group(1)))
    # "> X" or ">= X"
    m = re.match(r'^[>≥]=?\s*([\d.]+)', s)
    if m:
        return (float(m.group(1)), None)
    # "X - Y" or "X to Y"
    m = re.match(r'^([\d.]+)\s*[-–to]+\s*([\d.]+)', s)
    if m:
        return (float(m.group(1)), float(m.group(2)))
    return (None, None)


def classify_severity(value_str: str, normal_range: str) -> str:
    """
    Classify a lab value as normal / minor / major / critical.
    Returns 'normal' if value is within range or cannot be parsed.
    """
    if not value_str or not normal_range:
        return "normal"

    # Handle absent/present markers
    nr_lower = normal_range.strip().lower()
    if nr_lower in ("absent", "clear", "pale yellow", "negative"):
        val_lower = str(value_str).strip().lower()
        if val_lower in ("absent", "negative", "clear", "pale yellow", "not detected", "nil"):
            return "normal"
        return "minor"

    val = _extract_number(value_str)
    if val is None:
        return "normal"

    lower, upper = _parse_range(normal_range)
    if lower is None and upper is None:
        return "normal"

    # Calculate deviation ratio
    if lower is not None and upper is not None:
        # Two-sided range
        span = upper - lower
        if span <= 0:
            return "normal"
        if lower <= val <= upper:
            return "normal"
        deviation = max(lower - val, val - upper)
        ratio = deviation / span
        if ratio <= 0.3:
            return "minor"
        elif ratio <= 1.0:
            return "major"
        else:
            return "critical"

    elif upper is not None:
        # Upper-bound only ("< X")
        if val <= upper:
            return "normal"
        ratio = (val - upper) / upper if upper != 0 else 1
        if ratio <= 0.2:
            return "minor"
        elif ratio <= 0.75:
            return "major"
        else:
            return "critical"

    else:
        # Lower-bound only ("> X")
        if val >= lower:
            return "normal"
        ratio = (lower - val) / lower if lower != 0 else 1
        if ratio <= 0.2:
            return "minor"
        elif ratio <= 0.5:
            return "major"
        else:
            return "critical"


def parse_excel_lab_results(file_bytes: bytes) -> list[dict]:
    """Parse uploaded Excel and return classified findings."""
    import io
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes))
    ws = wb.active

    results = []
    # Find header row
    header_row = None
    col_map = {}
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
        if row and any(str(c or '').strip().lower() in ('test name', 'marker', 'test') for c in row):
            header_row = i
            for j, cell in enumerate(row):
                if cell:
                    key = str(cell).strip().lower()
                    col_map[key] = j
            break

    if header_row is None:
        header_row = 1
        col_map = {
            'test name': 0, 'what marker is this': 1, 'your value': 2,
            'normal range': 3, 'status': 4
        }

    name_col = col_map.get('test name', 0)
    value_col = col_map.get('your value', 2)
    range_col = col_map.get('normal range', 3)
    desc_col = col_map.get('what marker is this', 1)

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[name_col]:
            continue
        name = str(row[name_col]).strip()
        value = str(row[value_col] or '').strip() if len(row) > value_col else ''
        normal_range = str(row[range_col] or '').strip() if len(row) > range_col else ''
        description = str(row[desc_col] or '').strip() if len(row) > desc_col else ''

        if not name or not value or value in ('', 'None', '-', 'N/A'):
            continue

        # Find matching master marker for organ/test_type info
        master = next((m for m in MARKERS if m['name'].lower() == name.lower()), None)
        severity = classify_severity(value, normal_range)

        results.append({
            'name': name,
            'description': description or (master['description'] if master else ''),
            'value': value,
            'normal_range': normal_range,
            'unit': master['unit'] if master else '',
            'test_type': master['test_type'] if master else 'blood_urine',
            'organs': master['organs'] if master else [],
            'severity': severity,
        })

    return results


def generate_template_excel(patient=None, section=None) -> bytes:
    """Generate a downloadable Excel template with all markers pre-filled.

    If `patient` is provided, a patient-info banner is added at the top
    (name, Zen ID, Booking ID, scan date, age, gender) so the filled
    template is unambiguously tied to one patient.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Lab Results"

    # Header style
    header_fill = PatternFill(start_color="0E5C45", end_color="0E5C45", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["Test Name", "What Marker Is This", "Your Value", "Normal Range", "Unit", "Notes"]
    col_widths = [35, 35, 15, 20, 20, 40]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        ws.column_dimensions[chr(64 + col)].width = w

    # ── Patient banner ───────────────────────────────────────────────────────
    banner_offset = 0
    if patient:
        banner_fill = PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid")
        title_font = Font(bold=True, size=14, color="0E5C45")
        label_font = Font(bold=True, size=10, color="555555")
        value_font = Font(size=10, color="111111")

        # Row 1: title (merged)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        section_title = {
            "blood": " ZenLife — Blood Report Template",
            "urine": " ZenLife — Urine Analysis Template",
        }.get((section or "").lower(), " ZenLife — Lab Results Template")
        c = ws.cell(row=1, column=1, value=section_title)
        c.font = title_font
        c.fill = banner_fill
        c.alignment = Alignment(vertical='center')
        ws.row_dimensions[1].height = 26

        # Row 2: Patient name + Zen ID + Booking
        # We place them as label/value pairs across columns
        info_pairs = [
            ("Patient Name", patient.get("patient_name") or "—"),
            ("Zen ID", patient.get("zen_id") or "—"),
            ("Booking ID", patient.get("booking_id") or "—"),
        ]
        for i, (lab, val) in enumerate(info_pairs):
            lc = ws.cell(row=2, column=1 + i * 2, value=lab)
            vc = ws.cell(row=2, column=2 + i * 2, value=val)
            lc.font = label_font
            vc.font = value_font
            lc.fill = banner_fill
            vc.fill = banner_fill
            lc.alignment = Alignment(vertical='center')
            vc.alignment = Alignment(vertical='center')
        ws.row_dimensions[2].height = 20

        # Row 3: Age/Gender + Scan date
        scan_date = patient.get("scan_date") or "—"
        age_gender = ""
        age = patient.get("age")
        gender = patient.get("gender")
        if age and gender:
            age_gender = f"{age} yrs · {gender}"
        elif age:
            age_gender = f"{age} yrs"
        elif gender:
            age_gender = f"{gender}"
        else:
            age_gender = "—"
        info_pairs2 = [
            ("Age / Gender", age_gender),
            ("Scan Date", scan_date),
            ("", ""),
        ]
        for i, (lab, val) in enumerate(info_pairs2):
            lc = ws.cell(row=3, column=1 + i * 2, value=lab)
            vc = ws.cell(row=3, column=2 + i * 2, value=val)
            lc.font = label_font
            vc.font = value_font
            lc.fill = banner_fill
            vc.fill = banner_fill
            lc.alignment = Alignment(vertical='center')
            vc.alignment = Alignment(vertical='center')
        ws.row_dimensions[3].height = 20

        # Spacer row
        ws.row_dimensions[4].height = 8
        banner_offset = 4  # header row will start at row 5

    header_row = banner_offset + 1
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    ws.row_dimensions[header_row].height = 22

    # Group fills
    group_fills = {
        "Heart": PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
        "Endocrine & Metabolic Health": PatternFill(start_color="FFFBF0", end_color="FFFBF0", fill_type="solid"),
        "Liver & Digestive Health": PatternFill(start_color="F0FFF4", end_color="F0FFF4", fill_type="solid"),
        "Brain & Cognitive Health": PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid"),
        "Kidney & Urinary Health": PatternFill(start_color="FFF0FB", end_color="FFF0FB", fill_type="solid"),
        "Inflammation & Immune Health": PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid"),
        "General Health, Blood and Nutrients": PatternFill(start_color="F5F5F5", end_color="F5F5F5", fill_type="solid"),
        "Other": PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid"),
    }

    # When a specific blood/urine section is requested, drive the template
    # off SECTION_PARAMETERS so the parameter list matches the on-screen
    # admin panel exactly. Otherwise fall back to the broader MARKERS list.
    use_section_params = (section or "").lower() in ("blood", "urine")
    if use_section_params:
        from .section_params import SECTION_PARAMETERS
        section_defs = SECTION_PARAMETERS.get(section.lower(), [])
        # Wrap as the same shape as MARKERS so the loop below works
        markers_iter = [
            {
                "name": p["name"],
                "description": p.get("description", ""),
                "normal_range": p.get("normal", ""),
                "unit": p.get("unit", ""),
                "organs": [section.title()],
            }
            for p in section_defs
        ]
    else:
        markers_iter = MARKERS

    # Group markers by primary organ
    current_group = None
    row = header_row + 1
    for marker in markers_iter:
        primary_organ = marker['organs'][0] if marker['organs'] else 'Other'

        # Group header row
        if primary_organ != current_group:
            current_group = primary_organ
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            g_cell = ws.cell(row=row, column=1, value=f"  {primary_organ}")
            g_cell.font = Font(bold=True, size=10, color="444444")
            g_fill = group_fills.get(primary_organ, group_fills["Other"])
            for c in range(1, 7):
                ws.cell(row=row, column=c).fill = g_fill
            ws.row_dimensions[row].height = 16
            row += 1

        fill = group_fills.get(primary_organ, group_fills["Other"])
        values = [marker['name'], marker['description'], "", marker['normal_range'], marker['unit'], ""]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.border = border
            cell.alignment = Alignment(vertical='center')
            if col == 3:  # "Your Value" column — highlight
                cell.fill = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")

        ws.row_dimensions[row].height = 18
        row += 1

    # Instructions sheet
    ws2 = wb.create_sheet("Instructions")
    instructions = [
        ("ZenLife Lab Results Template", None),
        ("", None),
        ("HOW TO USE:", None),
        ("1. Fill in 'Your Value' column (column C) with the patient's lab result value.", None),
        ("2. Enter ONLY the numeric value — do not include units (e.g. enter 116, not '116 mg/dL').", None),
        ("3. Leave blank if a test was not performed.", None),
        ("4. Save as .xlsx and upload in the ZenLife Admin Panel.", None),
        ("", None),
        ("SEVERITY IS AUTO-CLASSIFIED:", None),
        ("• Normal   — Value within the stated normal range", None),
        ("• Minor    — Value up to 20-30% outside normal range", None),
        ("• Major    — Value 30-100% outside normal range", None),
        ("• Critical — Value more than 100% outside normal range", None),
    ]
    for i, (text, _) in enumerate(instructions, 1):
        cell = ws2.cell(row=i, column=1, value=text)
        if i == 1:
            cell.font = Font(bold=True, size=14, color="0E5C45")
        elif text.startswith("HOW TO USE") or text.startswith("SEVERITY"):
            cell.font = Font(bold=True, size=11)
    ws2.column_dimensions['A'].width = 80

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
